# openpyxl
# MSエクセルワークシートを操作するモジュール
# pip install openpyxl

# ワークシートのコピー
# ワークシートのリネーム
# セルのスタイル（枠線、フォント、色）設定
# 数式の編集
# 図の挿入
# セルのマージ
https://openpyxl.readthedocs.io/en/stable/


# 値の取得
import openpyxl
import pprint

def main():

    workbook = openpyxl.load_workbook("Some.xlsx")
    sheet    = workbook["Sheet1"]

    print(sheet["C5"].value)
    print(sheet["C6"].value)
    print(sheet["C7"].value)

    for i in sheet["C4:F8"][2]:
        print(i.value)

    rows = sheet.iter_rows(
        min_row=4, max_row=8, min_col=3, max_col=6)
    for i in list(rows)[0]:
        print(i.value)

    pprint.pprint(list(sheet.values))

    return

if __name__ == '__main__':
    main()



# 値の書き込み
import openpyxl

def main():

    workbook = openpyxl.load_workbook("Some.xlsx")
    sheet    = workbook["Sheet1"]

    sheet["C3"].value = "This is TEST Table."
    sheet.cell(row=2, column=3, value=100)

    workbook.save("Some.xlsx")

    return

if __name__ == '__main__':
    main()



